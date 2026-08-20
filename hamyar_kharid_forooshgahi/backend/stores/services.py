import csv
import io
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from django.db import transaction
from django.utils import timezone
from .models import Inventory, InventoryImportLog, Product, Store


class InventoryImportService:
    HEADERS = {'name', 'product_name'}

    @staticmethod
    def _normalize_row(row):
        normalized = {str(k).strip().lower(): v for k, v in row.items() if k is not None}
        name = normalized.get('name') or normalized.get('product_name') or normalized.get('product')
        category = normalized.get('category', '')
        price = normalized.get('price', 0)
        quantity = normalized.get('quantity')
        unit = normalized.get('unit', 'عدد')
        return {
            'name': str(name).strip() if name is not None else '',
            'category': str(category).strip(),
            'price': price,
            'quantity': quantity,
            'unit': str(unit).strip() or 'عدد',
        }

    @classmethod
    def parse_txt(cls, raw_bytes):
        text = raw_bytes.decode('utf-8-sig')
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if not lines:
            return []
        delimiter = ',' if ',' in lines[0] else '\t'
        reader = csv.DictReader(io.StringIO('\n'.join(lines)), delimiter=delimiter)
        return [cls._normalize_row(row) for row in reader]

    @classmethod
    def parse_excel(cls, uploaded_file):
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []
        header_names = [str(h).strip() if h is not None else '' for h in headers]
        return [cls._normalize_row(dict(zip(header_names, row))) for row in rows]

    @classmethod
    @transaction.atomic
    def import_rows(cls, store: Store, rows, source, filename=''):
        processed = 0
        errors = []
        for index, raw in enumerate(rows, start=2):
            row = cls._normalize_row(raw)
            try:
                if not row['name']:
                    raise ValueError('نام کالا الزامی است.')
                if row['quantity'] is None:
                    raise ValueError('موجودی الزامی است.')
                quantity = int(row['quantity'])
                if quantity < 0:
                    raise ValueError('موجودی نمی‌تواند منفی باشد.')
                price = Decimal(str(row['price']))
                if price < 0:
                    raise ValueError('قیمت نمی‌تواند منفی باشد.')
                category = row['category'] or 'سایر'
                product, _ = Product.objects.get_or_create(
                    store=store,
                    name=row['name'],
                    defaults={'category': category, 'price': price, 'unit': row['unit']},
                )
                changed = False
                if product.category != category:
                    product.category = category; changed = True
                if product.price != price:
                    product.price = price; changed = True
                if product.unit != row['unit']:
                    product.unit = row['unit']; changed = True
                if changed:
                    product.save(update_fields=['category', 'price', 'unit', 'updated_at'])
                Inventory.objects.update_or_create(product=product, store=store, defaults={'quantity': quantity})
                processed += 1
            except (ValueError, TypeError, InvalidOperation) as exc:
                errors.append({'row': index, 'error': str(exc)})
        log = InventoryImportLog.objects.create(
            store=store,
            source=source,
            file_name=filename,
            processed_rows=processed,
            failed_rows=len(errors),
        )
        return log, errors
